from flask import Flask, request, send_file, jsonify
from flask_cors import CORS
import openpyxl
import io
import os
import re
from datetime import datetime

app = Flask(__name__, static_folder='static', static_url_path='')
CORS(app)

PLANTILLA = os.path.join(os.path.dirname(__file__), 'plantilla.xlsx')

ITEMS_BIEN = [
    40,41,42,43,44,45,46,47,48,49,50,
    53,54,55,56,57,58,59,
    63,64,65,66,67,68,69,70,71,72,73,74,75,
    78,79,80,81,82,83,84,85,86,87,88,
    91,92,93,94,95,
    98,99,100,101,102,103,104,105,106,107,
    110,111,112,113,114,
    117,118,119,120,121,122,123,124,125,126,127,
]

FILA_LECTORES = 98

MESES = {
    '01':'ENERO','02':'FEBRERO','03':'MARZO','04':'ABRIL',
    '05':'MAYO','06':'JUNIO','07':'JULIO','08':'AGOSTO',
    '09':'SEPTIEMBRE','10':'OCTUBRE','11':'NOVIEMBRE','12':'DICIEMBRE'
}

def slug(text):
    text = text.upper()
    for src, dst in [('Á','A'),('À','A'),('É','E'),('È','E'),('Í','I'),('Ì','I'),
                     ('Ó','O'),('Ò','O'),('Ú','U'),('Ù','U'),('Ü','U'),('Ñ','N'),
                     ('á','A'),('é','E'),('í','I'),('ó','O'),('ú','U'),('ñ','N')]:
        text = text.replace(src, dst)
    text = re.sub(r'[^A-Z0-9]+', '_', text)
    return text.strip('_')

@app.route('/')
def index():
    return app.send_static_file('index.html')

@app.route('/generar', methods=['POST'])
def generar():
    data = request.json
    num_tienda = data.get('num_tienda', '').strip()
    direccion  = data.get('direccion', '').strip()
    fecha      = data.get('fecha', '').strip()
    tecnico    = data.get('tecnico', '').strip()
    sin_lector = data.get('sin_lector', False)

    if not num_tienda or not direccion or not fecha or not tecnico:
        return jsonify({'error': 'Faltan campos obligatorios'}), 400

    wb = openpyxl.load_workbook(PLANTILLA)
    ws = wb.active

    ws['C7'] = f'LIDL SUPERMERCADO S.A.U.  {num_tienda}'
    ws['C8'] = fecha
    ws['C8'].number_format = '@'
    ws['C9'] = direccion.upper()
    ws['C10'] = tecnico.upper()

    for fila in ITEMS_BIEN:
        if fila == FILA_LECTORES and sin_lector:
            ws.cell(row=fila, column=5).value = 'X'
        else:
            ws.cell(row=fila, column=3).value = 'X'

    # Nombre del archivo: Checklist_DIRECCION_NUMTIENDA_LIDL_MES_AÑO
    try:
        partes = fecha.split('/')
        mes = MESES.get(partes[1], partes[1])
        anio = partes[2]
    except:
        mes = 'FECHA'
        anio = ''

    dir_slug = slug(direccion)
    nombre = f'Checklist_{dir_slug}_{num_tienda}_LIDL_{mes}_{anio}.xlsx'

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=nombre
    )

if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5050))
    app.run(host='0.0.0.0', port=port)
